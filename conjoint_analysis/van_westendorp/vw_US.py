"""
Van Westendorp PSM — US Only
============================
Reads only the first 4 columns (US market) from the raw data file.
No currency conversion applied — values remain in USD.

Column order within those 4 columns:
  [Too Expensive, Expensive, Cheap, Too Cheap]

Outputs:
  - Console:          4 VW price intersection points
  - vw_us_output.xlsx Chart-ready cumulative % table + Price Points summary
  - vw_us_chart.png   PSM plot
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

# ─────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────
INPUT_FILE   = r"C:\Users\bbarber\OneDrive - BGB Group\Documents\1) Project work\BD\BD-014\Conjoint Analysis\17 June\vw_inputs.xlsx"   # accepts .xlsx or .csv
OUTPUT_EXCEL = r"C:\Users\bbarber\OneDrive - BGB Group\Documents\1) Project work\BD\BD-014\Conjoint Analysis\17 June\vw_us_output.xlsx"
OUTPUT_CHART = r"C:\Users\bbarber\OneDrive - BGB Group\Documents\1) Project work\BD\BD-014\Conjoint Analysis\17 June\vw_us_chart.png"
CHART_TITLE  = "Van Westendorp PSM — United States (USD)"
CURRENCY_SYM = "$"

# Column order within the 4 US columns (0-based offsets within the block)
QUESTIONS    = ["Too Expensive", "Expensive", "Cheap", "Too Cheap"]


# ─────────────────────────────────────────────
#  STEP 1 — Load raw data (skip 2 header rows, US cols only)
# ─────────────────────────────────────────────
def load_data(path: str) -> pd.DataFrame:
    print(f"Loading: {path}  [US only — columns 1–4]")
    if path.endswith(".csv"):
        raw = pd.read_csv(path, header=None, skiprows=2)
    else:
        raw = pd.read_excel(path, header=None, skiprows=2)

    raw = raw.iloc[:, :4].copy()          # US = first 4 columns
    raw = raw.apply(pd.to_numeric, errors="coerce")
    raw = raw.round(0)

    print(f"  Respondents loaded: {len(raw)}")
    return raw


# ─────────────────────────────────────────────
#  STEP 2 — Map columns to questions
# ─────────────────────────────────────────────
def stack_questions(df: pd.DataFrame) -> dict[str, np.ndarray]:
    stacked = {}
    for q_idx, question in enumerate(QUESTIONS):
        col = df.iloc[:, q_idx].dropna().astype(int)
        stacked[question] = col.values
        print(f"  {question}: {len(col)} responses, {col.nunique()} unique values")
    return stacked


# ─────────────────────────────────────────────
#  STEP 3 — Cumulative frequency curves
# ─────────────────────────────────────────────
def build_curves(
    stacked: dict[str, np.ndarray],
    prices: np.ndarray
) -> dict[str, np.ndarray]:
    def desc(data, p): return np.array([np.mean(data >= px) * 100 for px in p])
    def asc(data, p):  return np.array([np.mean(data <= px) * 100 for px in p])

    return {
        "Too Cheap":     desc(stacked["Too Cheap"],     prices),
        "Cheap":         desc(stacked["Cheap"],         prices),
        "Expensive":     asc(stacked["Expensive"],      prices),
        "Too Expensive": asc(stacked["Too Expensive"],  prices),
    }


# ─────────────────────────────────────────────
#  STEP 4 — Intersection finder
# ─────────────────────────────────────────────
def find_intersection(
    prices: np.ndarray, y1: np.ndarray, y2: np.ndarray
) -> tuple[float | None, float | None]:
    diff = y1 - y2
    sign_changes = np.where(np.diff(np.sign(diff)))[0]
    if not len(sign_changes):
        return None, None
    idx = sign_changes[0]
    x0, x1 = prices[idx], prices[idx + 1]
    d0, d1 = diff[idx], diff[idx + 1]
    x_int = x0 - d0 * (x1 - x0) / (d1 - d0)
    y_int = float(np.interp(x_int, prices, y1))
    return round(float(x_int), 1), round(y_int, 1)


# ─────────────────────────────────────────────
#  STEP 5 — Excel output
# ─────────────────────────────────────────────
def export_excel(
    prices: np.ndarray,
    curves: dict[str, np.ndarray],
    intersections: dict[str, tuple],
    path: str,
):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    plot_col_order = ["Too Cheap", "Cheap", "Expensive", "Too Expensive"]
    chart_df = pd.DataFrame({"Price": prices.astype(int)})
    for q in plot_col_order:
        chart_df[q] = np.round(curves[q], 2)

    int_labels = {
        "OPP": "Optimal Price Point (OPP)",
        "IDP": "Indifference Price Point (IDP)",
        "PMC": "Point of Marginal Cheapness (PMC)",
        "PME": "Point of Marginal Expensiveness (PME)",
    }
    summary_rows = [
        {"Price Point": int_labels[k], "Value": f"{CURRENCY_SYM}{round(v[0])}" if v[0] else "n/a"}
        for k, v in intersections.items()
    ]
    pmc, pme = intersections["PMC"][0], intersections["PME"][0]
    if pmc and pme:
        summary_rows.append({"Price Point": "Acceptable Price Range",
                              "Value": f"{CURRENCY_SYM}{round(pmc)} — {CURRENCY_SYM}{round(pme)}"})

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(chart_df).to_excel(writer, sheet_name="Chart Data",    index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Price Points", index=False)

    wb = load_workbook(path)
    ws = wb["Chart Data"]

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = bdr
            if cell.column > 1:
                cell.number_format = '0.00"%"'

    ws.column_dimensions["A"].width = 10
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 16

    n_rows = len(chart_df)
    chart = LineChart()
    chart.title  = CHART_TITLE
    chart.style  = 10
    chart.y_axis.title = "Cumulative %"
    chart.x_axis.title = f"Price ({CURRENCY_SYM})"
    chart.height, chart.width = 14, 24

    colors_hex = ["2563EB", "16A34A", "EA580C", "DC2626"]
    for i, (q, color) in enumerate(zip(plot_col_order, colors_hex), start=2):
        data_ref = Reference(ws, min_col=i, min_row=1, max_row=n_rows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[i - 2].graphicalProperties.line.solidFill = color
        chart.series[i - 2].graphicalProperties.line.width = 20000

    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1))
    ws.add_chart(chart, "G2")

    ws2 = wb["Price Points"]
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 20

    wb.save(path)
    print(f"\n  Excel saved → {path}")


# ─────────────────────────────────────────────
#  STEP 6 — Plot
# ─────────────────────────────────────────────
def plot_psm(
    prices: np.ndarray,
    curves: dict[str, np.ndarray],
    intersections: dict[str, tuple],
    path: str,
):
    PALETTE  = {"Too Cheap": "#2563EB", "Cheap": "#16A34A",
                "Expensive": "#EA580C", "Too Expensive": "#DC2626"}
    ICOLORS  = {"OPP": "#7C3AED", "IDP": "#0891B2",
                "PMC": "#059669", "PME": "#DB2777"}

    fig, ax = plt.subplots(figsize=(13, 7))
    fig.patch.set_facecolor("#F8F9FA")
    ax.set_facecolor("#FFFFFF")

    pmc_x, pme_x = intersections["PMC"][0], intersections["PME"][0]
    if pmc_x and pme_x:
        ax.axvspan(pmc_x, pme_x, alpha=0.07, color="#16A34A")

    for label, y in curves.items():
        ax.plot(prices, y, label=label, color=PALETTE[label], linewidth=2.2, zorder=3)

    for key, (x, y) in intersections.items():
        if x is None:
            continue
        color = ICOLORS[key]
        ax.axvline(x=x, color=color, linestyle="--", alpha=0.45, linewidth=1.2)
        ax.scatter([x], [y], color=color, s=90, zorder=6, edgecolors="white", linewidth=1.5)
        ax.annotate(f"  {key}\n  {CURRENCY_SYM}{x:,.0f}", xy=(x, y),
                    xytext=(0, 12), textcoords="offset points",
                    fontsize=8.5, fontweight="bold", color=color, va="bottom", ha="center")

    ax.set_xlabel(f"Price ({CURRENCY_SYM})", fontsize=12, labelpad=8)
    ax.set_ylabel("Cumulative %", fontsize=12, labelpad=8)
    ax.set_title(CHART_TITLE, fontsize=15, fontweight="bold", pad=14)
    ax.set_ylim(0, 102)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}%"))
    ax.grid(True, alpha=0.25, linestyle="--")
    ax.spines[["top", "right"]].set_visible(False)

    handles, labels = ax.get_legend_handles_labels()
    leg1 = ax.legend(handles, labels, loc="center right", fontsize=10,
                     framealpha=0.9, title="Curves")

    from matplotlib.lines import Line2D
    int_handles = [
        Line2D([0], [0], marker="o", color="w", markerfacecolor=ICOLORS[k],
               markersize=8, label=f"{k}: {CURRENCY_SYM}{v[0]:,.0f}" if v[0] else f"{k}: n/a")
        for k, v in intersections.items()
    ]
    ax.legend(handles=int_handles, loc="upper left", fontsize=9,
              framealpha=0.9, title="Key Price Points")
    ax.add_artist(leg1)

    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"  Chart saved → {path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    raw = load_data(INPUT_FILE)

    print("\nMapping questions...")
    stacked = stack_questions(raw)

    all_prices = np.sort(np.unique(np.concatenate(list(stacked.values()))))
    curves = build_curves(stacked, all_prices)

    intersections = {
        "OPP": find_intersection(all_prices, curves["Too Cheap"],  curves["Too Expensive"]),
        "IDP": find_intersection(all_prices, curves["Cheap"],      curves["Expensive"]),
        "PMC": find_intersection(all_prices, curves["Too Cheap"],  curves["Expensive"]),
        "PME": find_intersection(all_prices, curves["Cheap"],      curves["Too Expensive"]),
    }

    print("\n" + "=" * 48)
    print("  VAN WESTENDORP PRICE POINTS — US")
    print("=" * 48)
    labels = {"OPP": "Optimal Price Point        ",
              "IDP": "Indifference Price Point   ",
              "PMC": "Point of Marginal Cheapness",
              "PME": "Point of Marginal Expense  "}
    for key, (x, _) in intersections.items():
        val = f"{CURRENCY_SYM}{x:,.0f}" if x is not None else "not found"
        print(f"  {labels[key]}  {val}")
    pmc_x, pme_x = intersections["PMC"][0], intersections["PME"][0]
    if pmc_x and pme_x:
        print(f"\n  Acceptable Price Range: {CURRENCY_SYM}{pmc_x:,.0f} — {CURRENCY_SYM}{pme_x:,.0f}")
    print("=" * 48)

    export_excel(all_prices, curves, intersections, OUTPUT_EXCEL)
    plot_psm(all_prices, curves, intersections, OUTPUT_CHART)


if __name__ == "__main__":
    main()
