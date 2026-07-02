"""
Van Westendorp Price Sensitivity Meter (PSM) — Multi-Market Analysis
=====================================================================
Input:  Excel/CSV with 24 columns of raw USD data, 2 header rows
        Column order: US (0-3) | UK (4-7) | FR (8-11) | IT (12-15) | ES (16-19) | DE (20-23)
        Each block of 4 = [Too Cheap, Cheap, Expensive, Too Expensive]

Currency conversions applied:
  - UK  (cols 4–7):   USD → GBP  ×1.36612
  - DE  (cols 20–23): USD → EUR  ×1.173709

  NOTE: If FR, IT, ES data also needs USD→EUR conversion, update the
        EUR_COLS slice below to `slice(8, 24)` instead of `slice(20, 24)`.

Outputs:
  - Console:       4 VW price intersection points
  - vw_output.xlsx Unique sorted values for each of the 4 questions
  - vw_chart.png   PSM plot with intersection annotations
"""

import sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.patches import FancyArrowPatch

# ─────────────────────────────────────────────
#  CONFIGURATION — edit these as needed
# ─────────────────────────────────────────────
INPUT_FILE   = r"C:\Users\bbarber\OneDrive - BGB Group\Documents\1) Project work\BD\BD-014\Conjoint Analysis\17 June\vw_inputs.xlsx"   # accepts .xlsx or .csv
OUTPUT_EXCEL = r"C:\Users\bbarber\OneDrive - BGB Group\Documents\1) Project work\BD\BD-014\Conjoint Analysis\17 June\vw_output.xlsx"
OUTPUT_CHART = r"C:\Users\bbarber\OneDrive - BGB Group\Documents\1) Project work\BD\BD-014\Conjoint Analysis\17 June\vw_chart.png"


USD_TO_GBP   = 1.36612
USD_TO_EUR   = 1.173709

# Column slices (0-based, within the 24-column block)
GBP_COLS     = slice(4,  8)   # UK
EUR_COLS     = slice(20, 24)  # DE only — change to slice(8, 24) for FR+IT+ES+DE

QUESTIONS    = ["Too Expensive", "Expensive", "Cheap", "Too Cheap"]
MARKET_STARTS = [0, 4, 8, 12, 16, 20]  # first column of each of the 6 markets


# ─────────────────────────────────────────────
#  STEP 1 — Load raw data (skip 2 header rows)
# ─────────────────────────────────────────────
def load_data(path: str) -> pd.DataFrame:
    print(f"Loading: {path}")
    if path.endswith(".csv"):
        raw = pd.read_csv(path, header=None, skiprows=2)
    else:
        raw = pd.read_excel(path, header=None, skiprows=2)

    raw = raw.iloc[:, :24].copy()
    raw = raw.apply(pd.to_numeric, errors="coerce")

    if raw.shape[1] != 24:
        raise ValueError(f"Expected 24 data columns, found {raw.shape[1]}")

    print(f"  Rows loaded (respondents): {len(raw)}")
    return raw


# ─────────────────────────────────────────────
#  STEP 2 — Currency conversion
# ─────────────────────────────────────────────
def convert_currencies(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.iloc[:, GBP_COLS] = df.iloc[:, GBP_COLS].multiply(USD_TO_GBP)
    df.iloc[:, EUR_COLS] = df.iloc[:, EUR_COLS].multiply(USD_TO_EUR)
    df = df.round(0)
    print("  Currency conversion applied.")
    return df


# ─────────────────────────────────────────────
#  STEP 3 — Stack 6 markets into 4 question columns
# ─────────────────────────────────────────────
def stack_markets(df: pd.DataFrame) -> dict[str, np.ndarray]:
    """Returns dict: question → sorted array of unique integer values."""
    stacked = {}
    for q_idx, question in enumerate(QUESTIONS):
        cols = [start + q_idx for start in MARKET_STARTS]
        combined = pd.concat(
            [df.iloc[:, c] for c in cols], ignore_index=True
        ).dropna().astype(int)
        stacked[question] = combined.values
        print(f"  {question}: {len(combined)} total responses, "
              f"{combined.nunique()} unique values")
    return stacked


# ─────────────────────────────────────────────
#  STEP 4 — Cumulative frequency curves
# ─────────────────────────────────────────────
def build_curves(
    stacked: dict[str, np.ndarray],
    prices: np.ndarray
) -> dict[str, np.ndarray]:
    """
    Too Cheap & Cheap   → descending:  % of respondents where answer >= price
    Expensive & Too Exp → ascending:   % of respondents where answer <= price
    """
    n = {q: len(v) for q, v in stacked.items()}

    def desc(data, p): return np.array([np.mean(data >= px) * 100 for px in p])
    def asc(data, p):  return np.array([np.mean(data <= px) * 100 for px in p])

    return {
        "Too Cheap":     desc(stacked["Too Cheap"],     prices),
        "Cheap":         desc(stacked["Cheap"],         prices),
        "Expensive":     asc(stacked["Expensive"],      prices),
        "Too Expensive": asc(stacked["Too Expensive"],  prices),
    }


# ─────────────────────────────────────────────
#  STEP 5 — Intersection finder
# ─────────────────────────────────────────────
def find_intersection(
    prices: np.ndarray,
    y1: np.ndarray,
    y2: np.ndarray
) -> tuple[float | None, float | None]:
    diff = y1 - y2
    sign_changes = np.where(np.diff(np.sign(diff)))[0]
    if not len(sign_changes):
        return None, None
    idx = sign_changes[0]
    x0, x1 = prices[idx], prices[idx + 1]
    d0, d1 = diff[idx], diff[idx + 1]
    x_int = x0 - d0 * (x1 - x0) / (d1 - d0)
    y_int = float(np.interp(x_int, prices, y1))
    return round(float(x_int), 1), round(y_int, 1)


# ─────────────────────────────────────────────
#  STEP 6 — Excel output
# ─────────────────────────────────────────────
def export_excel(
    stacked: dict[str, np.ndarray],
    prices: np.ndarray,
    curves: dict[str, np.ndarray],
    intersections: dict[str, tuple],
    path: str,
):
    """
    Writes two sheets to the output Excel file:

    'Chart Data'   — one row per price point with cumulative % for all 4 curves.
                     Select columns A–E → Insert → Line Chart → done.

    'Price Points' — the 4 VW intersection values in a clean summary table.
    An embedded line chart is also added to the Chart Data sheet automatically.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    # ── Sheet 1: cumulative frequency table ──────────────────────────────
    plot_col_order = ["Too Cheap", "Cheap", "Expensive", "Too Expensive"]
    chart_df = pd.DataFrame({"Price": prices.astype(int)})
    for q in plot_col_order:
        chart_df[q] = np.round(curves[q], 2)

    # ── Sheet 2: price point summary ─────────────────────────────────────
    int_labels = {
        "OPP": "Optimal Price Point (OPP)",
        "IDP": "Indifference Price Point (IDP)",
        "PMC": "Point of Marginal Cheapness (PMC)",
        "PME": "Point of Marginal Expensiveness (PME)",
    }
    summary_rows = [
        {"Price Point": int_labels[k], "Value": round(v[0]) if v[0] else "n/a"}
        for k, v in intersections.items()
    ]
    pmc = intersections["PMC"][0]
    pme = intersections["PME"][0]
    if pmc and pme:
        summary_rows.append({"Price Point": "Acceptable Price Range",
                              "Value": f"{round(pmc)} — {round(pme)}"})
    summary_df = pd.DataFrame(summary_rows)

    # Write both sheets
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        chart_df.to_excel(writer, sheet_name="Chart Data",    index=False)
        summary_df.to_excel(writer, sheet_name="Price Points", index=False)

    # ── Style & embed chart via openpyxl ─────────────────────────────────
    wb = load_workbook(path)

    # Style Chart Data sheet
    ws = wb["Chart Data"]
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    thin  = Side(style="thin", color="CCCCCC")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = bdr
            if cell.column > 1:
                cell.number_format = '0.00"%"'

    ws.column_dimensions["A"].width = 10
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 16

    # Embed a line chart
    n_rows = len(chart_df)
    chart = LineChart()
    chart.title  = "Van Westendorp Price Sensitivity Meter"
    chart.style  = 10
    chart.y_axis.title = "Cumulative %"
    chart.x_axis.title = "Price"
    chart.height = 14
    chart.width  = 24

    colors_hex = ["2563EB", "16A34A", "EA580C", "DC2626"]
    for i, (q, color) in enumerate(zip(plot_col_order, colors_hex), start=2):
        data_ref = Reference(ws, min_col=i, min_row=1, max_row=n_rows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[i - 2].graphicalProperties.line.solidFill = color
        chart.series[i - 2].graphicalProperties.line.width = 20000

    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    chart.set_categories(cats)
    ws.add_chart(chart, "G2")

    # Style Price Points sheet
    ws2 = wb["Price Points"]
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 20

    wb.save(path)
    print(f"\n  Excel saved → {path}  (sheets: 'Chart Data' + 'Price Points')")


# ─────────────────────────────────────────────
#  STEP 7 — Plot
# ─────────────────────────────────────────────
def plot_psm(
    prices: np.ndarray,
    curves: dict[str, np.ndarray],
    intersections: dict[str, tuple],
    path: str,
):
    PALETTE = {
        "Too Cheap":     "#2563EB",   # blue
        "Cheap":         "#16A34A",   # green
        "Expensive":     "#EA580C",   # orange
        "Too Expensive": "#DC2626",   # red
    }
    ICOLORS = {
        "OPP": "#7C3AED",
        "IDP": "#0891B2",
        "PMC": "#059669",
        "PME": "#DB2777",
    }
    ILABELS = {
        "OPP": "Optimal Price Point",
        "IDP": "Indifference Price Point",
        "PMC": "Point of Marginal Cheapness",
        "PME": "Point of Marginal Expensiveness",
    }

    fig, ax = plt.subplots(figsize=(13, 7))
    fig.patch.set_facecolor("#F8F9FA")
    ax.set_facecolor("#FFFFFF")

    # Draw acceptable range band
    pmc_x = intersections["PMC"][0]
    pme_x = intersections["PME"][0]
    if pmc_x and pme_x:
        ax.axvspan(pmc_x, pme_x, alpha=0.07, color="#16A34A", label="_range")

    # Main curves
    for label, y in curves.items():
        ax.plot(prices, y, label=label, color=PALETTE[label],
                linewidth=2.2, zorder=3)

    # Intersection markers
    for key, (x, y) in intersections.items():
        if x is None:
            continue
        color = ICOLORS[key]
        ax.axvline(x=x, color=color, linestyle="--", alpha=0.45, linewidth=1.2)
        ax.scatter([x], [y], color=color, s=90, zorder=6, edgecolors="white",
                   linewidth=1.5)
        ax.annotate(
            f"  {key}\n  {x:,.0f}",
            xy=(x, y),
            xytext=(0, 12),
            textcoords="offset points",
            fontsize=8.5,
            fontweight="bold",
            color=color,
            va="bottom",
            ha="center",
        )

    # Axes & labels
    ax.set_xlabel("Price", fontsize=12, labelpad=8)
    ax.set_ylabel("Cumulative %", fontsize=12, labelpad=8)
    ax.set_title(
        "Van Westendorp Price Sensitivity Meter",
        fontsize=15,
        fontweight="bold",
        pad=14,
    )
    ax.set_ylim(0, 102)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}%"))
    ax.grid(True, alpha=0.25, linestyle="--")
    ax.spines[["top", "right"]].set_visible(False)

    # Legend — curves
    handles, labels = ax.get_legend_handles_labels()
    leg1 = ax.legend(handles, labels, loc="center right",
                     fontsize=10, framealpha=0.9, title="Curves")

    # Legend — intersections
    from matplotlib.lines import Line2D
    int_handles = [
        Line2D([0], [0], marker="o", color="w",
               markerfacecolor=ICOLORS[k], markersize=8,
               label=f"{k}: {v[0]:,.0f}" if v[0] else f"{k}: n/a")
        for k, v in intersections.items()
    ]
    ax.legend(handles=int_handles, loc="upper left",
              fontsize=9, framealpha=0.9, title="Key Price Points")
    ax.add_artist(leg1)

    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"  Chart saved → {path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    # 1. Load
    raw = load_data(INPUT_FILE)

    # 2. Convert currencies
    converted = convert_currencies(raw)

    # 3. Stack into 4 question columns
    print("\nStacking markets into 4 question columns...")
    stacked = stack_markets(converted)

    # 4. Build price range + cumulative curves
    all_prices = np.sort(
        np.unique(np.concatenate([v for v in stacked.values()]))
    )
    curves = build_curves(stacked, all_prices)

    # 5. Find intersections
    intersections = {
        "OPP": find_intersection(all_prices, curves["Too Cheap"],  curves["Too Expensive"]),
        "IDP": find_intersection(all_prices, curves["Cheap"],      curves["Expensive"]),
        "PMC": find_intersection(all_prices, curves["Too Cheap"],  curves["Expensive"]),
        "PME": find_intersection(all_prices, curves["Cheap"],      curves["Too Expensive"]),
    }

    print("\n" + "=" * 48)
    print("  VAN WESTENDORP PRICE POINTS")
    print("=" * 48)
    labels = {
        "OPP": "Optimal Price Point        ",
        "IDP": "Indifference Price Point   ",
        "PMC": "Point of Marginal Cheapness",
        "PME": "Point of Marginal Expense  ",
    }
    for key, (x, _) in intersections.items():
        val = f"{x:,.0f}" if x is not None else "not found"
        print(f"  {labels[key]}  {val}")
    pmc_x = intersections["PMC"][0]
    pme_x = intersections["PME"][0]
    if pmc_x and pme_x:
        print(f"\n  Acceptable Price Range:   {pmc_x:,.0f} — {pme_x:,.0f}")
    print("=" * 48)

    # 6. Export Excel
    export_excel(stacked, all_prices, curves, intersections, OUTPUT_EXCEL)

    # 7. Plot
    plot_psm(all_prices, curves, intersections, OUTPUT_CHART)


if __name__ == "__main__":
    main()