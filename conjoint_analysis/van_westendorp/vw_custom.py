"""
Van Westendorp PSM — Respondent Segment Filter
===============================================
Reads all 24 VW columns plus a 25th column containing a respondent
segment code (e.g. 1 = Clinicians, 2 = Clinical Leaders).

At runtime the script lists all available codes and their counts,
then asks you to enter the code you want to analyse. Only rows
matching that code are used for the analysis.

Column layout (first 24):
  US (0–3) | UK (4–7) | FR (8–11) | IT (12–15) | ES (16–19) | DE (20–23)
  Each block of 4: [Too Expensive, Expensive, Cheap, Too Cheap]
Column 25 (index 24): Respondent segment code

Currency conversions applied (same as main script):
  - UK  (cols 4–7):   USD → GBP  ×1.36612
  - DE  (cols 20–23): USD → EUR  ×1.173709

  NOTE: If FR, IT, ES also need USD→EUR, change
        EUR_COLS to slice(8, 24) in the CONFIGURATION block.

Outputs (filename includes the selected segment code):
  - Console:                      4 VW price intersection points
  - vw_segment_<code>_output.xlsx Chart-ready table + Price Points summary
  - vw_segment_<code>_chart.png   PSM plot labelled with the segment
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

# ─────────────────────────────────────────────
#  OUTPUT FILE LOCATIONS
#  All output files are saved to the SAME FOLDER
#  as this script. If you want them saved elsewhere,
#  edit the paths below to include a full folder path.
#
#  Example (Windows):
#    OUTPUT_EXCEL = r"C:\Users\YourName\Documents\vw_results\vw_segment_<code>_<label>_output.xlsx"
#    OUTPUT_CHART = r"C:\Users\YourName\Documents\vw_results\vw_segment_<code>_<label>_chart.png"
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────
INPUT_FILE   = r"C:\Users\bbarber\OneDrive - BGB Group\Documents\1) Project work\BD\BD-014\Data analysis\27 May - VW\VW_raw_data_roles.xlsx"   # accepts .xlsx or .csv

USD_TO_GBP  = 1.36612
USD_TO_EUR  = 1.173709

GBP_COLS    = slice(4,  8)    # UK
EUR_COLS    = slice(20, 24)   # DE only — change to slice(8, 24) for FR+IT+ES+DE

QUESTIONS   = ["Too Expensive", "Expensive", "Cheap", "Too Cheap"]
MARKET_STARTS = [0, 4, 8, 12, 16, 20]

# Optional: map code numbers to human-readable labels for chart titles.
# Add your own codes and names here, or leave empty ({}) to use the number only.
SEGMENT_LABELS = {
    1: "Clinicians",
    2: "Clinical Leaders",
    # 3: "Other Stakeholders",  ← add more as needed
}


# ─────────────────────────────────────────────
#  STEP 1 — Load raw data (skip 2 header rows)
# ─────────────────────────────────────────────
def load_data(path: str) -> pd.DataFrame:
    print(f"Loading: {path}")
    if path.endswith(".csv"):
        raw = pd.read_csv(path, header=None, skiprows=2)
    else:
        raw = pd.read_excel(path, header=None, skiprows=2)

    if raw.shape[1] < 25:
        raise ValueError(
            f"Expected at least 25 columns (1 segment code + 24 VW), "
            f"found {raw.shape[1]}. Please check your input file."
        )

    raw = raw.iloc[:, :25].copy()
    raw = raw.apply(pd.to_numeric, errors="coerce")
    print(f"  Total respondents loaded: {len(raw)}")
    return raw


# ─────────────────────────────────────────────
#  STEP 2 — Show available codes and prompt
# ─────────────────────────────────────────────
def select_segment(df: pd.DataFrame) -> tuple[int, pd.DataFrame]:
    segment_col = df.iloc[:, 24]
    counts = segment_col.value_counts().sort_index()

    print("\n" + "=" * 48)
    print("  AVAILABLE RESPONDENT SEGMENTS")
    print("=" * 48)
    print(f"  {'Code':<8} {'Label':<25} {'n':>6}")
    print(f"  {'-'*8} {'-'*25} {'-'*6}")
    for code, n in counts.items():
        label = SEGMENT_LABELS.get(int(code), "(no label — add to SEGMENT_LABELS)")
        print(f"  {int(code):<8} {label:<25} {n:>6}")
    print("=" * 48)

    while True:
        try:
            choice = int(input("\nEnter the segment code to analyse: ").strip())
            if choice not in counts.index.astype(int).tolist():
                print(f"  ✗ Code {choice} not found. Please choose from the list above.")
                continue
            break
        except ValueError:
            print("  ✗ Please enter a whole number.")

    filtered = df[segment_col == choice].copy()
    label = SEGMENT_LABELS.get(choice, f"Segment {choice}")
    print(f"\n  Selected: {label}  (n={len(filtered)} respondents)")
    return choice, filtered


# ─────────────────────────────────────────────
#  STEP 3 — Currency conversion
# ─────────────────────────────────────────────
def convert_currencies(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.iloc[:, GBP_COLS] = df.iloc[:, GBP_COLS].multiply(USD_TO_GBP)
    df.iloc[:, EUR_COLS] = df.iloc[:, EUR_COLS].multiply(USD_TO_EUR)
    df = df.round(0)
    print("  Currency conversion applied.")
    return df


# ─────────────────────────────────────────────
#  STEP 4 — Stack 6 markets into 4 question columns
# ─────────────────────────────────────────────
def stack_markets(df: pd.DataFrame) -> dict[str, np.ndarray]:
    stacked = {}
    for q_idx, question in enumerate(QUESTIONS):
        cols = [start + q_idx for start in MARKET_STARTS]
        combined = pd.concat(
            [df.iloc[:, c] for c in cols], ignore_index=True
        ).dropna().astype(int)
        stacked[question] = combined.values
        print(f"  {question}: {len(combined)} responses, "
              f"{combined.nunique()} unique values")
    return stacked


# ─────────────────────────────────────────────
#  STEP 5 — Cumulative frequency curves
# ─────────────────────────────────────────────
def build_curves(
    stacked: dict[str, np.ndarray],
    prices: np.ndarray
) -> dict[str, np.ndarray]:
    def desc(data, p): return np.array([np.mean(data >= px) * 100 for px in p])
    def asc(data, p):  return np.array([np.mean(data <= px) * 100 for px in p])

    return {
        "Too Cheap":     desc(stacked["Too Cheap"],     prices),
        "Cheap":         desc(stacked["Cheap"],         prices),
        "Expensive":     asc(stacked["Expensive"],      prices),
        "Too Expensive": asc(stacked["Too Expensive"],  prices),
    }


# ─────────────────────────────────────────────
#  STEP 6 — Intersection finder
# ─────────────────────────────────────────────
def find_intersection(
    prices: np.ndarray, y1: np.ndarray, y2: np.ndarray
) -> tuple[float | None, float | None]:
    diff = y1 - y2
    sign_changes = np.where(np.diff(np.sign(diff)))[0]
    if not len(sign_changes):
        return None, None
    idx = sign_changes[0]
    x0, x1 = prices[idx], prices[idx + 1]
    d0, d1 = diff[idx], diff[idx + 1]
    x_int = x0 - d0 * (x1 - x0) / (d1 - d0)
    y_int = float(np.interp(x_int, prices, y1))
    return round(float(x_int), 1), round(y_int, 1)


# ─────────────────────────────────────────────
#  STEP 7 — Excel output
# ─────────────────────────────────────────────
def export_excel(
    prices: np.ndarray,
    curves: dict[str, np.ndarray],
    intersections: dict[str, tuple],
    segment_label: str,
    path: str,
):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    plot_col_order = ["Too Cheap", "Cheap", "Expensive", "Too Expensive"]
    chart_df = pd.DataFrame({"Price": prices.astype(int)})
    for q in plot_col_order:
        chart_df[q] = np.round(curves[q], 2)

    int_labels = {
        "OPP": "Optimal Price Point (OPP)",
        "IDP": "Indifference Price Point (IDP)",
        "PMC": "Point of Marginal Cheapness (PMC)",
        "PME": "Point of Marginal Expensiveness (PME)",
    }
    summary_rows = [
        {"Price Point": int_labels[k], "Value": round(v[0]) if v[0] else "n/a"}
        for k, v in intersections.items()
    ]
    pmc, pme = intersections["PMC"][0], intersections["PME"][0]
    if pmc and pme:
        summary_rows.append({"Price Point": "Acceptable Price Range",
                              "Value": f"{round(pmc)} — {round(pme)}"})

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(chart_df).to_excel(writer, sheet_name="Chart Data",    index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Price Points", index=False)

    wb = load_workbook(path)
    ws = wb["Chart Data"]

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = bdr
            if cell.column > 1:
                cell.number_format = '0.00"%"'

    ws.column_dimensions["A"].width = 10
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 16

    n_rows = len(chart_df)
    chart_title = f"Van Westendorp PSM — {segment_label}"
    chart = LineChart()
    chart.title  = chart_title
    chart.style  = 10
    chart.y_axis.title = "Cumulative %"
    chart.x_axis.title = "Price"
    chart.height, chart.width = 14, 24

    colors_hex = ["2563EB", "16A34A", "EA580C", "DC2626"]
    for i, (q, color) in enumerate(zip(plot_col_order, colors_hex), start=2):
        data_ref = Reference(ws, min_col=i, min_row=1, max_row=n_rows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[i - 2].graphicalProperties.line.solidFill = color
        chart.series[i - 2].graphicalProperties.line.width = 20000

    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1))
    ws.add_chart(chart, "G2")

    ws2 = wb["Price Points"]
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 20

    wb.save(path)
    print(f"\n  Excel saved → {path}")


# ─────────────────────────────────────────────
#  STEP 8 — Plot
# ─────────────────────────────────────────────
def plot_psm(
    prices: np.ndarray,
    curves: dict[str, np.ndarray],
    intersections: dict[str, tuple],
    segment_label: str,
    path: str,
):
    PALETTE = {"Too Cheap": "#2563EB", "Cheap": "#16A34A",
               "Expensive": "#EA580C", "Too Expensive": "#DC2626"}
    ICOLORS = {"OPP": "#7C3AED", "IDP": "#0891B2",
               "PMC": "#059669", "PME": "#DB2777"}

    fig, ax = plt.subplots(figsize=(13, 7))
    fig.patch.set_facecolor("#F8F9FA")
    ax.set_facecolor("#FFFFFF")

    pmc_x, pme_x = intersections["PMC"][0], intersections["PME"][0]
    if pmc_x and pme_x:
        ax.axvspan(pmc_x, pme_x, alpha=0.07, color="#16A34A")

    for label, y in curves.items():
        ax.plot(prices, y, label=label, color=PALETTE[label], linewidth=2.2, zorder=3)

    for key, (x, y) in intersections.items():
        if x is None:
            continue
        color = ICOLORS[key]
        ax.axvline(x=x, color=color, linestyle="--", alpha=0.45, linewidth=1.2)
        ax.scatter([x], [y], color=color, s=90, zorder=6, edgecolors="white", linewidth=1.5)
        ax.annotate(f"  {key}\n  {x:,.0f}", xy=(x, y),
                    xytext=(0, 12), textcoords="offset points",
                    fontsize=8.5, fontweight="bold", color=color, va="bottom", ha="center")

    ax.set_xlabel("Price", fontsize=12, labelpad=8)
    ax.set_ylabel("Cumulative %", fontsize=12, labelpad=8)
    ax.set_title(f"Van Westendorp PSM — {segment_label}",
                 fontsize=15, fontweight="bold", pad=14)
    ax.set_ylim(0, 102)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}%"))
    ax.grid(True, alpha=0.25, linestyle="--")
    ax.spines[["top", "right"]].set_visible(False)

    handles, labels = ax.get_legend_handles_labels()
    leg1 = ax.legend(handles, labels, loc="center right", fontsize=10,
                     framealpha=0.9, title="Curves")

    from matplotlib.lines import Line2D
    int_handles = [
        Line2D([0], [0], marker="o", color="w", markerfacecolor=ICOLORS[k],
               markersize=8, label=f"{k}: {v[0]:,.0f}" if v[0] else f"{k}: n/a")
        for k, v in intersections.items()
    ]
    ax.legend(handles=int_handles, loc="upper left", fontsize=9,
              framealpha=0.9, title="Key Price Points")
    ax.add_artist(leg1)

    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"  Chart saved → {path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    # 1. Load (all 25 columns)
    # Derive output folder from input file location
    input_dir = os.path.dirname(os.path.abspath(INPUT_FILE))
    print(f"  Outputs will be saved to: {input_dir}")
    raw = load_data(INPUT_FILE)

    # 2. Prompt user to select a segment
    code, filtered = select_segment(raw)
    segment_label = SEGMENT_LABELS.get(code, f"Segment {code}")

    # 3. Currency conversion (on VW columns only; ignore col 25)
    converted = convert_currencies(filtered)

    # 4. Stack markets
    print("\nStacking markets into 4 question columns...")
    stacked = stack_markets(converted)

    # 5. Build curves
    all_prices = np.sort(np.unique(np.concatenate(list(stacked.values()))))
    curves = build_curves(stacked, all_prices)

    # 6. Find intersections
    intersections = {
        "OPP": find_intersection(all_prices, curves["Too Cheap"],  curves["Too Expensive"]),
        "IDP": find_intersection(all_prices, curves["Cheap"],      curves["Expensive"]),
        "PMC": find_intersection(all_prices, curves["Too Cheap"],  curves["Expensive"]),
        "PME": find_intersection(all_prices, curves["Cheap"],      curves["Too Expensive"]),
    }

    print("\n" + "=" * 48)
    print(f"  VAN WESTENDORP PRICE POINTS — {segment_label.upper()}")
    print("=" * 48)
    label_map = {"OPP": "Optimal Price Point        ",
                 "IDP": "Indifference Price Point   ",
                 "PMC": "Point of Marginal Cheapness",
                 "PME": "Point of Marginal Expense  "}
    for key, (x, _) in intersections.items():
        val = f"{x:,.0f}" if x is not None else "not found"
        print(f"  {label_map[key]}  {val}")
    pmc_x, pme_x = intersections["PMC"][0], intersections["PME"][0]
    if pmc_x and pme_x:
        print(f"\n  Acceptable Price Range:   {pmc_x:,.0f} — {pme_x:,.0f}")
    print("=" * 48)

    # 7. Output filenames include the segment code
    safe_label = segment_label.replace(" ", "_")
    out_excel = os.path.join(input_dir, f"vw_segment_{code}_{safe_label}_output.xlsx")
    out_chart = os.path.join(input_dir, f"vw_segment_{code}_{safe_label}_chart.png")

    export_excel(all_prices, curves, intersections, segment_label, out_excel)
    plot_psm(all_prices, curves, intersections, segment_label, out_chart)


if __name__ == "__main__":
    main()
